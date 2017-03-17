from openpyxl import load_workbook

wb = load_workbook('db.xlsx')
sheet = wb.create_sheet()
print(wb.get_sheet_names())
